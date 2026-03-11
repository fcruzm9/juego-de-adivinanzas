# ============================================================
# JUEGO: ADIVINA EL NÚMERO
# Autor: Felipe Cruz
# ============================================================
# PAQUETES NECESARIOS:
#   - openpyxl  --> pip install openpyxl   (leer y escribir Excel)
#   - matplotlib --> pip install matplotlib (mostrar gráficos)
#   - random    --> ya incluido en Python
#   - os        --> ya incluido en Python
#
# FICHEROS QUE CREA EL PROGRAMA:
#   - Estadisticas.xlsx
#   - Ruta Windows : C:\EjerciciosPython\Estadisticas.xlsx
#   - Ruta Mac/Linux: ~/EjerciciosPython/Estadisticas.xlsx
# ============================================================

import random
import os
import openpyxl
import matplotlib.pyplot as plt


# ------------------------------------------------------------
# FUNCIONES DE MENÚ
# ------------------------------------------------------------

def mostrar_menu():
    print("\n========================================")
    print("    BIENVENIDO A 'ADIVINA EL NÚMERO'")
    print("========================================")
    print("1. Partida modo solitario")
    print("2. Partida 2 jugadores")
    print("3. Estadística")
    print("4. Salir")
    print("========================================")


def mostrar_dificultad():
    print("\n--- Elige la dificultad ---")
    print("1. Fácil  (20 intentos)")
    print("2. Medio  (12 intentos)")
    print("3. Difícil ( 5 intentos)")


# ------------------------------------------------------------
# FUNCIONES DE VALIDACIÓN
# ------------------------------------------------------------

def validar_opcion(minimo, maximo):
    while True:
        try:
            opcion = int(input("Elige una opción: "))
            if minimo <= opcion <= maximo:
                return opcion
            else:
                print("Opción no válida, elige entre", minimo, "y", maximo)
        except ValueError:
            print("Debes escribir un número entero")


def validar_numero(minimo, maximo, mensaje):
    while True:
        try:
            numero = int(input(mensaje))
            if minimo <= numero <= maximo:
                return numero
            else:
                print("El número debe estar entre", minimo, "y", maximo)
        except ValueError:
            print("Debes escribir un número entero")


def pedir_nombre():
    while True:
        nombre = input("Escribe tu nombre y apellido: ").strip()
        if nombre != "":
            return nombre
        print("El nombre no puede estar vacío")


# ------------------------------------------------------------
# FUNCIONES DEL EXCEL (módulo propio de estadísticas)
# ------------------------------------------------------------

def obtener_ruta():
    if os.name == "nt":
        carpeta = r"C:\EjerciciosPython"
    else:
        carpeta = os.path.join(os.path.expanduser("~"), "EjerciciosPython")
    os.makedirs(carpeta, exist_ok=True)
    return os.path.join(carpeta, "Estadisticas.xlsx")


def crear_excel_si_no_existe(ruta):
    if not os.path.exists(ruta):
        libro = openpyxl.Workbook()
        hoja = libro.active
        hoja.title = "Partidas"
        hoja.append(["Jugador", "Modo", "Dificultad", "Resultado", "Intentos"])
        libro.save(ruta)


def guardar_partida(nombre, modo, dificultad, resultado, intentos):
    ruta = obtener_ruta()
    crear_excel_si_no_existe(ruta)

    if modo == 1:
        nombre_modo = "Solitario"
    else:
        nombre_modo = "2 Jugadores"

    if dificultad == 1:
        nombre_dif = "Fácil"
    elif dificultad == 2:
        nombre_dif = "Medio"
    else:
        nombre_dif = "Difícil"

    libro = openpyxl.load_workbook(ruta)
    hoja = libro.active
    hoja.append([nombre, nombre_modo, nombre_dif, resultado, intentos])
    libro.save(ruta)


def mostrar_estadisticas():
    ruta = obtener_ruta()
    crear_excel_si_no_existe(ruta)

    libro = openpyxl.load_workbook(ruta)
    hoja = libro.active

    filas = list(hoja.iter_rows(values_only=True))
    datos = filas[1:]  # quitar la cabecera

    if len(datos) == 0:
        print("\nNo hay partidas registradas todavía")
        return

    # Mostrar tabla en consola
    print("\n" + "=" * 60)
    print(f"{'JUGADOR':<20} {'MODO':<14} {'DIFICULTAD':<10} {'RESULTADO':<10} INTENTOS")
    print("=" * 60)
    for fila in datos:
        print(f"{str(fila[0]):<20} {str(fila[1]):<14} {str(fila[2]):<10} {str(fila[3]):<10} {fila[4]}")
    print("=" * 60)

    # Contar victorias y derrotas
    gano = 0
    perdio = 0
    for fila in datos:
        if fila[3] == "Ganó":
            gano += 1
        elif fila[3] == "Perdió":
            perdio += 1

    # Contar por dificultad
    facil = 0
    medio = 0
    dificil = 0
    for fila in datos:
        if fila[2] == "Fácil":
            facil += 1
        elif fila[2] == "Medio":
            medio += 1
        elif fila[2] == "Difícil":
            dificil += 1

    # Mostrar gráficos
    fig, ejes = plt.subplots(1, 2, figsize=(10, 5))
    fig.suptitle("Estadísticas - Adivina el Número", fontsize=13)

    ejes[0].bar(["Ganaron", "Perdieron"], [gano, perdio], color=["green", "red"])
    ejes[0].set_title("Resultados globales")
    ejes[0].set_ylabel("Partidas")

    ejes[1].bar(["Fácil", "Medio", "Difícil"], [facil, medio, dificil], color=["blue", "orange", "red"])
    ejes[1].set_title("Partidas por dificultad")
    ejes[1].set_ylabel("Partidas")

    plt.tight_layout()
    plt.show()


# ------------------------------------------------------------
# FUNCIÓN PRINCIPAL DE JUEGO
# ------------------------------------------------------------

def jugar(modo, dificultad):
    if dificultad == 1:
        intentos_max = 20
    elif dificultad == 2:
        intentos_max = 12
    else:
        intentos_max = 5

    if modo == 1:
        numero_secreto = random.randint(1, 1000)
        print("\nEl ordenador ha elegido un número entre 1 y 1000")
        print("Tienes", intentos_max, "intentos. ¡Buena suerte!\n")
    else:
        numero_secreto = validar_numero(1, 1000, "Jugador 1, escribe el número secreto (1-1000): ")
        os.system("cls" if os.name == "nt" else "clear")
        input("Jugador 2, presiona Enter cuando estés listo...")
        print("\nTienes", intentos_max, "intentos. ¡Adivina el número!\n")

    intentos_usados = 0

    while intentos_usados < intentos_max:
        restantes = intentos_max - intentos_usados
        print("Intentos restantes:", restantes)
        intento = validar_numero(1, 1000, "Adivina el número: ")
        intentos_usados += 1

        if intento == numero_secreto:
            print("\n¡CORRECTO! Adivinaste en", intentos_usados, "intento(s)")
            nombre = pedir_nombre()
            input("Presiona Enter para volver al menú...")
            return nombre, "Ganó", intentos_usados
        elif intento < numero_secreto:
            print("El número secreto es MAYOR\n")
        else:
            print("El número secreto es MENOR\n")

    print("\n¡Sin intentos! El número era:", numero_secreto)
    nombre = pedir_nombre()
    input("Presiona Enter para volver al menú...")
    return nombre, "Perdió", intentos_usados


# ------------------------------------------------------------
# MENÚ PRINCIPAL
# ------------------------------------------------------------

def main():
    opcion = 0

    while opcion != 4:
        mostrar_menu()
        opcion = validar_opcion(1, 4)

        if opcion == 1 or opcion == 2:
            mostrar_dificultad()
            dificultad = validar_opcion(1, 3)
            os.system("cls" if os.name == "nt" else "clear")
            nombre, resultado, intentos = jugar(opcion, dificultad)
            guardar_partida(nombre, opcion, dificultad, resultado, intentos)

        elif opcion == 3:
            mostrar_estadisticas()
            input("\nPresiona Enter para volver al menú...")

        elif opcion == 4:
            print("\n¡Hasta la próxima!\n")


main()
